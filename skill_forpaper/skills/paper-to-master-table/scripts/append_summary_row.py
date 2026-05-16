import argparse
import json
from typing import Any, Dict, List, Optional

import openpyxl


def _norm_header(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip().lower().replace(" ", "")


def _find_col(headers: List[Any], needles: List[str]) -> Optional[int]:
    norms = [_norm_header(h) for h in headers]
    for i, h in enumerate(norms, start=1):
        if not h:
            continue
        for n in needles:
            if n in h:
                return i
    return None


def _is_header_row_empty(headers: List[Any]) -> bool:
    return all(h is None or str(h).strip() == "" for h in headers)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to SUMMARY.xlsx")
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--data-json", help="Row data as JSON object")
    g.add_argument("--data-file", help="Path to a UTF-8 JSON file containing the row object")
    args = ap.parse_args()

    if args.data_file:
        with open(args.data_file, "r", encoding="utf-8") as f:
            data: Dict[str, Any] = json.load(f)
    else:
        data = json.loads(args.data_json)

    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb.worksheets[0]  # first sheet = master table

    # Read header row (row 1) up to max_column; ensure at least some columns exist.
    max_col = max(ws.max_column, 6)
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]

    desired = [
        ("title", "标题", ["标题", "title"]),
        ("authors_inst", "作者（机构/学校）", ["作者（机构/学校）", "作者", "authors", "机构", "institution"]),
        ("year", "发表年份", ["发表年份", "年份", "year", "date"]),
        ("venue", "期刊/会议", ["期刊/会议", "期刊", "会议", "venue", "journal", "conference"]),
        ("abstract", "摘要", ["摘要", "abstract"]),
        ("method_content", "研究方法&内容", ["研究方法&内容", "研究方法", "方法内容", "方法", "content", "method"]),
        ("experiments", "实验", ["实验", "experiment"]),
        ("related_work", "相关工作", ["相关工作", "related work"]),
        ("reflection", "自我思考", ["自我思考", "思考", "reflection"]),
        ("code", "有无代码", ["有无代码", "代码", "code", "github", "repository"]),
    ]

    if _is_header_row_empty(headers):
        for idx, (_, header_zh, _) in enumerate(desired, start=1):
            ws.cell(row=1, column=idx).value = header_zh
        max_col = len(desired)
        headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    else:
        # Ensure required columns exist; if missing, append new columns at the end.
        for _, header_zh, needles in desired:
            col = _find_col(headers, [n.lower() for n in needles])
            if col is None:
                max_col += 1
                ws.cell(row=1, column=max_col).value = header_zh
                headers.append(header_zh)

    # Build mapping from our logical fields -> column indices
    col_map: Dict[str, int] = {}
    for key, _, needles in desired:
        col = _find_col(headers, [n.lower() for n in needles])
        if col is None:
            # Fallback: append at end (shouldn't happen due to ensure step)
            max_col += 1
            ws.cell(row=1, column=max_col).value = key
            headers.append(key)
            col = max_col
        col_map[key] = col

    # Prepare an empty row with current width and fill known columns.
    row_values = [None] * len(headers)
    for k, v in data.items():
        if k not in col_map:
            continue
        row_values[col_map[k] - 1] = v

    ws.append(row_values)
    appended_row_idx = ws.max_row
    wb.save(args.xlsx)

    print(f"Appended to: {args.xlsx}")
    print(f"Sheet: {ws.title}")
    print(f"Row: {appended_row_idx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
