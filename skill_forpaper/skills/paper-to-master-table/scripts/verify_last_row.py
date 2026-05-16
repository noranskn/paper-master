import openpyxl


def main() -> None:
    xlsx = r"F:\论文\SUMMARY.xlsx"
    wb = openpyxl.load_workbook(xlsx)
    ws = wb.worksheets[0]  # first sheet
    r = ws.max_row
    max_col = ws.max_column
    print("max_col", max_col)
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    row = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
    print("max_row", r)
    code_str = "论文中无"
    contains_code_str = any(v == code_str for v in row)
    print("contains_code_str", contains_code_str)
    title_snippet = "A data analytics-driven approach"
    authors_snippet = "Prince Sultan University"
    title_cols = [i for i, v in enumerate(row, start=1) if isinstance(v, str) and title_snippet in v]
    authors_cols = [i for i, v in enumerate(row, start=1) if isinstance(v, str) and authors_snippet in v]
    print("title_cols", title_cols)
    print("authors_cols", authors_cols)
    def _count_newline(s: object) -> None:
        if not isinstance(s, str):
            return
        # check both newline types and literal backslash-n
        print(
            "  newline_presence:",
            "actual_LF(\\n)=" + str("\n" in s),
            "actual_CR(\\r)=" + str("\r" in s),
            "literal_backslash_n(\\\\n)=" + str("\\n" in s),
            "literal_backslash_r(\\\\r)=" + str("\\r" in s),
        )
        # print first few control characters (if any)
        ctrl_positions = [(i, ord(ch)) for i, ch in enumerate(s) if ord(ch) < 32 and ch not in ("\t",)]
        print("  control_char_positions_sample", ctrl_positions[:10])

    for c, (h, v) in enumerate(zip(headers, row), start=1):
        if h is None and v is None:
            continue
        if h in {"摘要", "实验", "相关工作", "自我思考", "标题", "作者（机构/学校）"}:
            print(f"  newline_check for {h!r}:")
            _count_newline(v)
        print(f"col {c} | header={repr(h)} | value={repr(v)}")


if __name__ == "__main__":
    main()

